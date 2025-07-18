import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from io import BytesIO
from PIL import Image
from datetime import datetime
import os
import re
from tqdm import tqdm
from urllib.parse import urlparse, parse_qs
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def ask_proxy():
    use_proxy = input("🔌 是否使用 proxy？(y/n): ").strip().lower()
    proxies = {}
    if use_proxy == 'y':
        proxy_addr = input("請輸入 proxy URL（例：http://127.0.0.1:7890）: ").strip()
        proxies = {"http": proxy_addr, "https": proxy_addr}
        print("✅ 已設定 proxy")
    return proxies


def fetch_target_url():
    return input("🔗 請輸入抓取目標網址（例：https://www.spec.org/...pattern=6787P）:\n").strip()


def fetch_result_list(url, proxies):
    print(f"📥 正在抓取列表頁面: {url}")
    resp = requests.get(url, proxies=proxies, verify=False)
    soup = BeautifulSoup(resp.text, "html.parser")
    rows = soup.find_all("tr")[1:]  # 略過表頭

    valid_rows = []
    for row in rows:
        cells = row.find_all("td")
        if len(cells) >= 9:
            links = cells[8].find_all("a", href=True)
            if any(a['href'].endswith(".txt") for a in links):
                valid_rows.append(row)

    print(f"🔍 共找到 {len(valid_rows)} 筆含 .txt 的結果")
    return valid_rows


def parse_txt_block(txt_content, sponser):
    lines = txt_content.splitlines()
    start = end = None
    for i, line in enumerate(lines):
        if "Benchmark Results Summary" in line:
            start = i
        elif "System Under Test" in line:
            end = i
            break
    if start is None or end is None:
        return []

    data_block = lines[start:end]
    extracted = []
    for line in data_block:
        line = line.strip()
        if re.match(r"^\d+%.*\|\s*\d", line):
            parts = [x.strip().replace(",", "") for x in line.split("|")]
            if len(parts) == 5:
                extracted.append(parts)
        elif "Active Idle" in line:
            parts = [x.strip().replace(",", "") for x in line.split("|")]
            extracted.append(["Active Idle", "", parts[1], parts[2], parts[3]])
        elif "sum of ssj_ops" in line:
            val = line.split("= |")[-1].strip().replace(",", "")
            extracted.append([sponser, "", "", "∑ssj_ops / ∑power =", val])
    return extracted


def extract_info(row, proxies):
    base_url = "https://www.spec.org"
    cells = row.find_all("td")
    sponsor = cells[0].get_text(strip=True)
    ssj_ops = cells[2].get_text(strip=True).replace(",", "")
    links = cells[8].find_all("a", href=True)

    txt_url = ""
    for a in links:
        if a["href"].endswith(".txt"):
            txt_url = base_url + a["href"]
            uid = re.search(r'power_ssj2008-\d+-\d+', a["href"])
            uid = uid.group(0) if uid else "graph"
            break
    if not txt_url:
        return None

    txt_resp = requests.get(txt_url, proxies=proxies, verify=False)
    content = txt_resp.text
    table_data = parse_txt_block(content, sponsor)

    jvm_match = re.search(r"JVM Command-line Options:\s+(.*?)(?=\n\s*JVM Affinity:|\n\s*\w)", content, re.DOTALL)
    jvm_opts = " ".join(line.strip() for line in jvm_match.group(1).splitlines()) if jvm_match else "Not found"

    boot_fw = re.search(r"Boot Firmware Settings\s*\n(.*?)(?=\n\s*Management Firmware Settings)", content, re.DOTALL)
    if boot_fw:
        lines = [line.strip() for line in boot_fw.group(1).splitlines() if line.strip()]
        lines.sort(key=lambda x: x.lower())
        boot_fw_str = "\n".join(lines)
    else:
        boot_fw_str = "Not found"

    return {
        "sponsor": sponsor,
        "ssj_ops": f"{int(ssj_ops):,} overall ssj_ops/watt",
        "txt_url": txt_url,
        "jvm_options": jvm_opts,
        "boot_fw_settings": boot_fw_str,
        "table_data": table_data,
        "img_url": txt_url.replace(".txt", ".png"),
        "uid": uid
    }


def save_to_excel(results, proxies, url):
    wb = Workbook()
    # Data Sheet（改為第一個 sheet）
    ws_data = wb.active
    ws_data.title = "Summary"
    ws_data.column_dimensions["K"].width = 100
    headers = ["Target Load", "Actual Load", "ssj_ops", "Avg Power (W)", "Perf/Watt"]
    row_data = 1
    temp_imgs = []
    image_start_col = 6  # H 欄，即右側

    print("\n🖼️ 開始載入圖片並插入 Excel...\n")
    for res in tqdm(results, desc="插入圖片", unit="圖表"):
        start_row = row_data
        for col_idx, h in enumerate(headers, start=1):
            ws_data.cell(row=row_data, column=col_idx, value=h)
        row_data += 1
        for line in res["table_data"]:
            for col_idx, val in enumerate(line, start=1):
                ws_data.cell(row=row_data, column=col_idx, value=val)
            row_data += 1

        boot_fw_lines = res["boot_fw_settings"].splitlines()
        end_row = row_data - 1
        if boot_fw_lines:
            cell_range = f"K{start_row}:K{end_row}"
            ws_data.merge_cells(cell_range)
            cell = ws_data[cell_range.split(":")[0]]
            cell.value = res["boot_fw_settings"]
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        try:
            img_data = requests.get(res["img_url"], proxies=proxies, verify=False).content
            img = Image.open(BytesIO(img_data))
            img_path = f"{res['uid']}.png"
            img.save(img_path)
            temp_imgs.append(img_path)

            excel_img = ExcelImage(img_path)
            excel_img.width, excel_img.height = 250, 250
            img_cell = f"{get_column_letter(image_start_col)}{start_row}"
            ws_data.add_image(excel_img, img_cell)

        except Exception as e:
            print(f"⚠️ 圖片載入失敗：{res['img_url']} - {e}")
            row_data += 3

    # Summary Sheet（改為第二個 sheet）
    ws_summary = wb.create_sheet(title="JVM_Options")
    headers = ["ssj_ops", "Tester", "URL", "JVM_Options"]
    for col_idx, h in enumerate(headers, start=1):
        ws_summary.cell(row=1, column=col_idx, value=h)

    row_sum = 2
    for res in results:
        ws_summary.cell(row=row_sum, column=1, value=res["ssj_ops"])
        ws_summary.cell(row=row_sum, column=2, value=res["sponsor"])
        ws_summary.cell(row=row_sum, column=3, value=res["txt_url"])
        ws_summary.cell(row=row_sum, column=4, value=res["jvm_options"])
        row_sum += 1

    # BIOS Sheet（第三個 sheet）
    ws_bios = wb.create_sheet(title="BIOS_Options")
    for col_idx, res in enumerate(results, start=1):
        col_letter = get_column_letter(col_idx)
        ws_bios.column_dimensions[col_letter].width = 60

        # ssj_ops (第1列)
        cell1 = ws_bios.cell(row=1, column=col_idx, value=res["ssj_ops"])
        cell1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # boot_fw_settings (第2列)
        cell2 = ws_bios.cell(row=2, column=col_idx, value=res["boot_fw_settings"])
        cell2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    parsed_url = urlparse(url)
    params = parse_qs(parsed_url.query)
    prefix = params.get("pattern", [None])[0] or params.get("crit-CPU", [None])[0] or "result"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_specpwr_results_{timestamp}.xlsx"
    wb.save(filename)
    print(f"\n✅ 已儲存至 {filename}")

    # 刪除暫存圖片
    for img_path in temp_imgs:
        try:
            os.remove(img_path)
            # print(f"🗑️ 刪除圖片: {img_path}")
        except Exception as e:
            print(f"⚠️ 刪除失敗: {img_path} - {e}")


def main():
    proxies = ask_proxy()
    url = fetch_target_url()
    rows = fetch_result_list(url, proxies)
    results = []

    print("\n🔄 開始處理 .txt 結果...\n")
    for row in tqdm(rows, desc="處理中", unit="項"):
        res = extract_info(row, proxies)
        if res:
            results.append(res)

    # 依 ssj_ops 數值排序（去除逗號與單位字串）
    results.sort(key=lambda x: int(x["ssj_ops"].split()[0].replace(",", "")), reverse=True)

    save_to_excel(results, proxies, url)


if __name__ == "__main__":
    main()
